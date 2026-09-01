#!/usr/bin/env python3
"""
build_outputs.py — Generate Excel + HTML Retail Business Dashboard from SQL CSV output.

Usage:
    python build_outputs.py <csv_path>
    python build_outputs.py <csv_path> --snapshot-date 2026-05-31
    python build_outputs.py <csv_path> --output-dir /custom/path/

The script:
  1. Loads the SQL output CSV (one row per week_start × dashboard_channel).
  2. Recomputes Total Retail = Amazon + Walmart + HD + Other Omni + TTS US for HTML
     (Total Retail in CSV may exclude TTS; HTML always includes it).
  3. Populates the Excel template (inputs/Dashboard_Template.xlsx),
     adding hardcoded per-channel values + formulas for Total Retail SUMs,
     % of Target, % of Total Retail YTD, Amazon Ads (On-site + DSP), and
     % of budget (month + QTD + YTD) rows.
  4. Adds a "Finance Monthly Actuals" placeholder sheet.
  5. Renders an HTML dashboard with sticky headers, color-coded %, and notes.

Outputs:
    outputs/<snapshot-date>/Retail Business Dashboard_<snapshot-date>.xlsx
    outputs/<snapshot-date>/Retail Business Dashboard_<snapshot-date>.html
"""

import argparse
import csv
import os
import re
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# Constants
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
INPUTS_DIR = PROJECT_ROOT / "inputs"
TEMPLATE_PATH = INPUTS_DIR / "Dashboard_Template.xlsx"
MARKETING_PLAN_PATH = INPUTS_DIR / "FY26_Marketing_Plan.xlsx"
PROMO_PLAN_PATH = INPUTS_DIR / "FY26_Promo_Plan.xlsx"
FINANCE_ACTUALS_PATH = INPUTS_DIR / "finance" / "Marketing_Spend_2026YTD.csv"
RESULTS_BASE = PROJECT_ROOT / "outputs"

# 2026 first Mon-Sun week starts here (Mon 2025-12-29 = ISO W1 2026)
WEEK_BASE = date(2025, 12, 29)
# N_WEEKS is computed dynamically from CSV at runtime — varies each refresh as
# more weeks become complete (CSV's WHERE filter drops future weeks).

# Channels summed into Total Retail (per user 2026-06-01: Amazon = 1P only,
# Amazon 3P excluded from Total Retail; TTS US renamed to 'TikTok Shop' and
# now includes MX)
SUM_CHANNELS = ['Amazon', 'Walmart', 'Home Depot', 'Other Omni', 'TikTok Shop']

# CSV fields summed into Total Retail (SUM-based; new_customers NOT summable)
SUM_FIELDS = [
    'SELL_IN_GROSS_REV_ACTUAL', 'SELL_IN_GROSS_REV_TARGET',
    'SELL_IN_GROSS_REV_CAM_ACTUAL', 'SELL_IN_GROSS_REV_NONCAM_ACTUAL',
    'SELL_THROUGH_UNITS_ACTUAL', 'SELL_THROUGH_UNITS_TARGET',
    'SELL_THROUGH_UNITS_CAM_ACTUAL', 'SELL_THROUGH_UNITS_NONCAM_ACTUAL',
    'SELL_THROUGH_UNITS_CAM_TARGET', 'SELL_THROUGH_UNITS_NONCAM_TARGET',
]

# Excel row map per channel section. Updated 2026-07-08 for the new template:
# (a) Total Retail NC block simplified — dropped single-week "As % of all units"
#     & single-week "% of Target"; added "% new customers - YTD" (R21 = SUM(NC)/
#     SUM(units)) and "% of Target - YTD" (R22 = R21 - R23, pp); Target R23 = 52%.
#     Everything from R21 down shifted up by 1 vs the old 212-row layout.
# (b) A-col of each Target/Budgeted row carries a template =SUM(E:BE) full-year
#     total (script never touches col A). B-col gained "Actual breakdown" /
#     "Real-time tracking" labels (labels only).
# Amazon NC block (R43-47) intentionally NOT simplified — fill actual (R43) + 52%
# target (R47) only; ratio rows R44/45/46 stay blank.
CH_ROWS = {
    'Amazon':      {'si_a': 29,  'si_t': 33,  'si_pct': 31,  'st_a': 35,  'st_t': 39,  'st_pct': 37,  'st_cam': 40,  'st_nc': 41,  'nc_a': 43,  'nc_t': 47, 'promo': 49, 'ads_on': 77, 'ads_dsp': 80, 'ads_tot': 83},
    'Walmart':     {'si_a': 91,  'si_t': 95,  'si_pct': 93,  'st_a': 97,  'st_t': 101, 'st_pct': 99,  'st_cam': 102, 'st_nc': 103, 'promo': 105, 'ads_tot': 119},
    'Home Depot':  {'si_a': 127, 'si_t': 131, 'si_pct': 129, 'st_a': 133, 'st_t': 137, 'st_pct': 135, 'st_cam': 138, 'st_nc': 139, 'promo': 141, 'ads_tot': 155},
    'Other Omni':  {'si_a': 163, 'si_t': 167, 'si_pct': 165, 'st_a': 169, 'st_t': 173, 'st_pct': 171, 'st_cam': 174, 'st_nc': 175, 'promo': 177},
    'TikTok Shop': {'si_a': 193, 'si_t': 197, 'si_pct': 195, 'st_a': 199, 'st_t': 203, 'st_pct': 201, 'st_cam': 204, 'st_nc': 205, 'promo': 207, 'ads_tot': 221},
}

# Total Retail rollup row positions. NC: nc_occ_ytd (R21 "% new customers - YTD" =
# SUM(NC)/SUM(units)), nc_pct (R22 "% of Target - YTD" pp = R21 - R23), nc_t (R23 = 52%).
TR_ROWS = {'si_a': 8, 'si_t': 11, 'si_pct': 9, 'st_a': 13, 'st_t': 16, 'st_pct': 14, 'st_cam': 17, 'st_nc': 18, 'nc_a': 20, 'nc_occ_ytd': 21, 'nc_pct': 22, 'nc_t': 23, 'nc_cam': 24, 'nc_nc': 25}

# "% of Total - YTD" rows: (pct_row, channel_data_row, all_channels_denom_row)
YTD_PCT_ROWS = [
    (30,  29,  3), (36,  35,  4),  # Amazon
    (92,  91,  3), (98,  97,  4),  # Walmart
    (128, 127, 3), (134, 133, 4),  # HD
    (164, 163, 3), (170, 169, 4),  # Other Omni
    (194, 193, 3), (200, 199, 4),  # TikTok Shop
]

# "% of Target - YTD" rows (sell-in + sell-through only): (pct_ytd_row, actual, target).
# Plain cumulative SUM/SUM (first-week target backfilled). NC handled in Pass 2b.
TARGET_YTD_PCT_ROWS = [
    (10,  8,   11),  (15,  13,  16),    # Total Retail si / st
    (32,  29,  33),  (38,  35,  39),    # Amazon       si / st
    (94,  91,  95),  (100, 97,  101),   # Walmart      si / st
    (130, 127, 131), (136, 133, 137),   # Home Depot   si / st
    (166, 163, 167), (172, 169, 173),   # Other Omni   si / st
    (196, 193, 197), (202, 199, 203),   # TikTok Shop  si / st
]

# Marketing blocks with a full "% of budget" group — Promo + Total Marketing +
# (2026-08-25) the Non-MDF / MDF split of Total Marketing, Amazon only.
# 6-tuple (actual, finance, budget, pct_month, pct_qtd, pct_ytd); all 3 % use finance.
# Non-MDF + MDF are a SECOND, independent breakdown of the same Total Marketing
# total (the other being Ads + Other Marketing below) — not additive with it.
MARKETING_BLOCKS = [
    (49,  50,  51,  52,  53,  54),    # Amazon Promo
    (56,  57,  58,  59,  60,  61),    # Amazon Total Marketing
    (63,  64,  65,  66,  67,  68),    # Amazon Non-MDF   (2026-08-25, new)
    (70,  71,  72,  73,  74,  75),    # Amazon MDF       (2026-08-25, new)
    (105, 106, 107, 108, 109, 110),   # Walmart Promo
    (112, 113, 114, 115, 116, 117),   # Walmart Total Marketing
    (141, 142, 143, 144, 145, 146),   # HD Promo
    (148, 149, 150, 151, 152, 153),   # HD Total Marketing
    (177, 178, 179, 180, 181, 182),   # Other Omni Promo
    (184, 185, 186, 187, 188, 189),   # Other Omni Total marketing
    (207, 208, 209, 210, 211, 212),   # TikTok Promo
    (214, 215, 216, 217, 218, 219),   # TikTok Total marketing
]

# dash_row -> budgeted_monthly row. Promo + Total Marketing + 2-row Ads/Other details.
PLAN_TARGETS = [
    # PROMO
    {'dash_row': 51,  'kind': 'promo',     'channel': 'Amazon 1P'},
    {'dash_row': 107, 'kind': 'promo',     'channel': 'Walmart'},
    {'dash_row': 143, 'kind': 'promo',     'channel': 'Home Depot'},
    {'dash_row': 179, 'kind': 'promo',     'channel': 'Best Buy'},     # Other Omni
    {'dash_row': 209, 'kind': 'promo',     'channel': 'TikTok Shop'},
    # TOTAL MARKETING (all ads + other-marketing buckets)
    {'dash_row': 58,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads', 'DSP', 'MDF', 'Influencer', 'Strategic Vendor Service - FTE', 'Prime Video Ads', 'Affiliate']},
    {'dash_row': 114, 'kind': 'marketing', 'channel': 'Walmart',    'buckets': ['Ad Spend', 'In-Store Display', 'Content']},
    {'dash_row': 150, 'kind': 'marketing', 'channel': 'Home Depot', 'buckets': ['Marketing package', 'In-Store Display', 'Events', 'Brand Advocate']},
    {'dash_row': 186, 'kind': 'marketing_multi_channel', 'channels': ['Best Buy', 'Costco', 'Other Retail']},
    {'dash_row': 216, 'kind': 'marketing', 'channel': 'TikTok Shop','buckets': ['Ad Spend', 'Affiliate Commisions', 'Free Samples', 'Livestreaming', 'Affiliate Maintenance and Tools', 'Video']},
    # AMAZON — Non-MDF / MDF split of Total Marketing (2026-08-25, new).
    # Budget-side MDF is a real, distinct bucket in the FY26 Marketing Tracker
    # (not derived) — Non-MDF is simply Total Marketing's 7 other buckets.
    {'dash_row': 65,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads', 'DSP', 'Influencer', 'Strategic Vendor Service - FTE', 'Prime Video Ads', 'Affiliate']},
    {'dash_row': 72,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['MDF']},
    # AMAZON detail (Ads / Other Marketing — unchanged breakdown, Other Marketing
    # still includes the 'MDF' bucket alongside it; two independent cuts, see above)
    {'dash_row': 78,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads']},
    {'dash_row': 81,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['DSP']},
    {'dash_row': 84,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads', 'DSP']},
    {'dash_row': 87,  'kind': 'marketing', 'channel': 'Amazon 1P',  'buckets': ['MDF', 'Influencer', 'Strategic Vendor Service - FTE', 'Prime Video Ads', 'Affiliate']},
    # WALMART detail
    {'dash_row': 120, 'kind': 'marketing', 'channel': 'Walmart',    'buckets': ['Ad Spend']},
    {'dash_row': 123, 'kind': 'marketing', 'channel': 'Walmart',    'buckets': ['In-Store Display', 'Content']},
    # HD detail
    {'dash_row': 156, 'kind': 'marketing', 'channel': 'Home Depot', 'buckets': ['Marketing package']},
    {'dash_row': 159, 'kind': 'marketing', 'channel': 'Home Depot', 'buckets': ['In-Store Display', 'Events', 'Brand Advocate']},
    # TIKTOK detail
    {'dash_row': 222, 'kind': 'marketing', 'channel': 'TikTok Shop','buckets': ['Ad Spend']},
    {'dash_row': 225, 'kind': 'marketing', 'channel': 'TikTok Shop','buckets': ['Affiliate Commisions', 'Free Samples', 'Livestreaming', 'Affiliate Maintenance and Tools', 'Video']},
]

# Other Marketing detail rows (2-row blocks) — gray-fill removal safety net.
OTHER_MARKETING_GRAY_ROWS = [86, 87, 122, 123, 158, 159, 224, 225]

# New-Customers target = flat 52% (constant). Total Retail R23, Amazon R47.
NC_TARGET_PCT = 0.52
NC_TARGET_ROWS = [23, 47]

# (FINANCE_MONTHLY_METRICS and add_finance_monthly_sheet() removed 2026-06-01.
# Finance Actuals rows now live inline in the Dashboard sheet, one per metric.)

# ── Finance Actuals (Monthly) source: inputs/finance/Marketing_Spend_2026YTD.csv ──
# Monthly NetSuite GL spend exposed by
# DATA_MART.RETAIL_DASHBOARD.MARKETING_SPEND_YTD as a long table:
# PERIOD, CATEGORY, CHANNEL, MDF_SPLIT, SPEND. The local finance_actuals.sql is
# a read-only wrapper around that view. Refreshed ~monthly in place.
# Replaces the old wide 'Channel × Category' xlsx.
#
# 3 categories (Discounts excluded in the query — never folded into Promo):
#   Promos       = accts 4094 + 4095 + 4098 (sign-flipped to positive in query)
#   Advertising  = acct  6601
#   Other Mktg   = accts 6102 + 6107 + 6108 + 6602 + 6604 + 6605 + 6610
# 5 channels, each → one dashboard 'Finance Actuals (Monthly)' row below:
#   Amazon 1P→Amazon, Walmart→Walmart, Home Depot→Home Depot,
#   Other Retail→Other Omni (BB + Costco + Other Retail), TTS→TikTok Shop.
#   HD CHANNEL includes Home Depot Canada; TTS includes TikTok Shop Mexico
#   (folded in by the query — they belong to their own sections, not Intl).
# CATEGORY + CHANNEL values line up directly with FINANCE_ROW_MAP below.
# Note worth remembering:
#   • HD ad spend is mostly booked under Other Mktg (Channel Marketing 6102),
#     NOT Advertising 6601 — so HD's finance 'Advertising' is tiny and will not
#     match the weekly HD_ADS_DAILY actual (different row anyway). Expected.
MONTH_ABBR = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
              'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}

# (finance_channel, finance_category) → dashboard Finance-Actuals row number.
# 2026-06-22: finance now lands in only TWO rows per channel — Promo and the
# combined **Total Marketing** row. Advertising + Other Mktg both point at the
# Total Marketing finance row (apply_finance_actuals ACCUMULATES, so they sum) —
# this sidesteps the Amazon-ads-booked-as-6102 classification problem entirely.
# Detail Ads / Other Marketing finance rows are intentionally left blank.
#
# 2026-08-25: Amazon MDF has no separate GL account — accounting tags it via
# memo ("paid through AVC account balance") on acct 6102 (Other Mktg). The CSV
# now carries a 4th column MDF_SPLIT ('Non-MDF' / 'MDF' / 'N/A - Promos');
# load_finance_actuals() additionally keys those rows by the 3-tuple
# (channel, category, mdf_split) WITHOUT changing the existing 2-tuple keys
# above (Non-MDF + MDF still roll up into the same Total Marketing total —
# see the 3 entries below). Non-MDF = Advertising's Non-MDF slice (all of it,
# MDF never hits 6601) + Other Mktg's Non-MDF slice.
FINANCE_ROW_MAP = {
    ('Amazon 1P',    'Promos'):      50,    # Amazon → Promo
    ('Amazon 1P',    'Advertising'): 57,    # Amazon → Total Marketing (Adv + Other summed)
    ('Amazon 1P',    'Other Mktg'):  57,    # Amazon → Total Marketing
    ('Amazon 1P', 'Advertising', 'Non-MDF'): 64,   # Amazon → Non-MDF (2026-08-25)
    ('Amazon 1P', 'Other Mktg',  'Non-MDF'): 64,   # Amazon → Non-MDF (2026-08-25)
    ('Amazon 1P', 'Other Mktg',  'MDF'):     71,   # Amazon → MDF     (2026-08-25)
    ('Walmart',      'Promos'):      106,   # Walmart → Promo
    ('Walmart',      'Advertising'): 113,   # Walmart → Total Marketing
    ('Walmart',      'Other Mktg'):  113,   # Walmart → Total Marketing
    ('Home Depot',   'Promos'):      142,   # Home Depot → Promo
    ('Home Depot',   'Advertising'): 149,   # Home Depot → Total Marketing
    ('Home Depot',   'Other Mktg'):  149,   # Home Depot → Total Marketing
    ('Other Retail', 'Promos'):      178,   # Other Omni → Promo
    ('Other Retail', 'Advertising'): 185,   # Other Omni → Total marketing
    ('Other Retail', 'Other Mktg'):  185,   # Other Omni → Total marketing
    ('TTS',          'Promos'):      208,   # TikTok Shop → Promo
    ('TTS',          'Advertising'): 215,   # TikTok Shop → Total marketing
    ('TTS',          'Other Mktg'):  215,   # TikTok Shop → Total marketing
}

# HTML equivalent: (dashboard_ch, metric_label) → (finance_channel, category).
FINANCE_HTML_MAP = {
    ('Amazon',      'Promo'):               ('Amazon 1P',    'Promos'),
    ('Amazon',      'Ads (On-site + DSP)'): ('Amazon 1P',    'Advertising'),
    ('Amazon',      'Other Marketing'):     ('Amazon 1P',    'Other Mktg'),
    ('Walmart',     'Promo'):           ('Walmart',      'Promos'),
    ('Walmart',     'Other Marketing'): ('Walmart',      'Other Mktg'),
    ('Home Depot',  'Promo'):           ('Home Depot',   'Promos'),
    ('Home Depot',  'Ads'):             ('Home Depot',   'Advertising'),
    ('Home Depot',  'Other Marketing'): ('Home Depot',   'Other Mktg'),
    ('Other Omni',  'Promo'):           ('Other Retail', 'Promos'),
    ('Other Omni',  'Total Marketing'): ('Other Retail', 'Other Mktg'),
    ('TikTok Shop', 'Promo'):           ('TTS', 'Promos'),
    ('TikTok Shop', 'Ads (GMV Max)'):   ('TTS', 'Advertising'),
    ('TikTok Shop', 'Other Marketing'): ('TTS', 'Other Mktg'),
}


# ============================================================
# Helpers
# ============================================================

def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safediv(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b


def col_letter(c):
    s = ''
    while c > 0:
        c, r = divmod(c - 1, 26)
        s = chr(65 + r) + s
    return s


def week_col(week_start_str):
    """Map a Mon-Sun week_start string → Excel column index (E=5, F=6, ...)."""
    d = date.fromisoformat(week_start_str)
    delta = (d - WEEK_BASE).days
    if delta < 0 or delta % 7 != 0:
        return None
    return 5 + delta // 7


def parse_snapshot_date_from_csv_path(csv_path):
    """Extract YYYY-MM-DD from CSV filename. Returns None if not found."""
    m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(csv_path))
    return m.group(1) if m else None


def last_mon_of_month_col(year, month):
    """Return Excel col index of the LAST Mon-Sun week whose Sunday is in the
    given (year, month) — i.e., the week whose end-of-week falls in this month.
    Examples: Jan 2026 → Mon 2026-01-19 (week ends Sun 2026-01-25);
              Feb 2026 → Mon 2026-02-16 (week ends Sun 2026-02-22);
              May 2026 → Mon 2026-05-25 (week ends Sun 2026-05-31).
    """
    # Last day of month
    if month == 12:
        next_month_first = date(year + 1, 1, 1)
    else:
        next_month_first = date(year, month + 1, 1)
    last_day = next_month_first - timedelta(days=1)
    # Roll back to nearest Sunday on or before last_day
    days_back_to_sun = (last_day.weekday() + 1) % 7  # Mon=0, Sun=6 → +1 mod 7 lands Mon=1, Sun=0
    last_sun = last_day - timedelta(days=days_back_to_sun)
    last_mon = last_sun - timedelta(days=6)
    delta = (last_mon - WEEK_BASE).days
    if delta < 0 or delta % 7 != 0:
        return None
    return 5 + delta // 7


def load_promo_plan():
    """Load FY26 Promo Tracker Summary → dict {channel: {(year, month): value_dollars}}.

    Channel rows recognized: Amazon 1P / Home Depot / Walmart / Best Buy / TikTok Shop.
    """
    if not MARKETING_PLAN_PATH.exists() or not PROMO_PLAN_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(PROMO_PLAN_PATH, data_only=True)
    ws = wb['Summary']
    # Header row 2: col B onwards = months (2026-01-01 ... 2026-12-01)
    month_cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(2, c).value
        if hasattr(v, 'year'):
            month_cols[c] = (v.year, v.month)
    result = {}
    for r in range(3, ws.max_row + 1):
        channel = ws.cell(r, 2).value
        if not channel or not isinstance(channel, str):
            continue
        channel = channel.strip()
        if channel not in ('Amazon 1P', 'Home Depot', 'Walmart', 'Best Buy', 'TikTok Shop'):
            continue
        result[channel] = {}
        for c, ym in month_cols.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):  # keep 0 values — they're intentional plan zeros
                result[channel][ym] = float(v)
    return result


def load_marketing_plan():
    """Load FY26 Marketing Tracker → dict {(channel, bucket): {(year, month): value_dollars}}."""
    if not MARKETING_PLAN_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(MARKETING_PLAN_PATH, data_only=True)
    ws = wb['Marketing_Tracker']
    # Header row 1: col E onwards = months
    month_cols = {}
    for c in range(5, ws.max_column + 1):
        v = ws.cell(1, c).value
        if hasattr(v, 'year'):
            month_cols[c] = (v.year, v.month)
    result = {}
    for r in range(2, ws.max_row + 1):
        channel = ws.cell(r, 2).value
        bucket = ws.cell(r, 3).value
        if not channel or not bucket:
            continue
        if str(bucket).strip().lower() == 'total':
            continue  # skip total rows; we sum buckets ourselves
        key = (str(channel).strip(), str(bucket).strip())
        result[key] = {}
        for c, ym in month_cols.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):  # keep 0s — intentional plan zeros (e.g. retail promo Jan/Feb)
                result[key][ym] = float(v)
    return result


def apply_plan_targets(ws, marketing_plan, promo_plan, max_col):
    """Write monthly plan values into dashboard 'Budgeted' rows.

    Placement: each (channel, month) value goes into the FIRST Mon-Sun week
    whose Monday is in that calendar month. Other weeks of that month: blank.

    Values divided by 1000 (Marketing/Promo plan files are in $; dashboard is $k).
    """
    n_written = 0
    for entry in PLAN_TARGETS:
        dash_row = entry['dash_row']
        kind = entry['kind']
        # Build {(year,month): total_dollars} for this dashboard row
        monthly = {}
        if kind == 'promo':
            ch = entry['channel']
            monthly = promo_plan.get(ch, {})
        elif kind == 'marketing':
            ch = entry['channel']
            for bucket in entry['buckets']:
                bucket_monthly = marketing_plan.get((ch, bucket), {})
                for ym, v in bucket_monthly.items():
                    monthly[ym] = monthly.get(ym, 0) + v
        elif kind == 'marketing_multi_channel':
            # Sum across multiple channels, all buckets within each
            for ch in entry['channels']:
                for (c, b), m in marketing_plan.items():
                    if c != ch:
                        continue
                    for ym, v in m.items():
                        monthly[ym] = monthly.get(ym, 0) + v
        # Write each monthly value into its first-Monday-of-month column
        for (year, month), val in monthly.items():
            col = last_mon_of_month_col(year, month)
            if col is None or col > max_col + 30:  # write up to ~7 months past visible data
                continue
            ws.cell(dash_row, col).value = val / 1000.0  # $ → $k
            n_written += 1
    return n_written


def load_finance_actuals():
    """Load Marketing_Spend_2026YTD.csv → {key: {(year,month): $}}.

    Reads the LONG NetSuite-derived table: PERIOD ('Mon YYYY'), CATEGORY
    ('Promos'/'Advertising'/'Other Mktg'), CHANNEL ('Amazon 1P'/'Walmart'/
    'Home Depot'/'Other Retail'/'TTS'), MDF_SPLIT ('Non-MDF'/'MDF'/
    'N/A - Promos', added 2026-08-25 — Amazon MDF has no separate GL account,
    so finance_actuals.sql tags it via memo instead), SPEND ($, Promo already
    positive, Discounts already excluded — see finance_actuals.sql).

    Every row is aggregated under its plain (channel, category) 2-tuple key —
    UNCHANGED behavior, MDF_SPLIT is invisible here, so Non-MDF + MDF still sum
    back into the old Other-Mktg-only total (backward compatible with
    FINANCE_ROW_MAP's existing 2-tuple entries / Total Marketing / HTML's
    finance_total_mktg). Rows also get a SECOND, granular (channel, category,
    mdf_split) 3-tuple key so the Non-MDF / MDF dashboard rows can be sourced
    without disturbing the 2-tuple totals. Older CSVs without the MDF_SPLIT
    column simply never populate the 3-tuple keys (loop below no-ops).

    Returns {} if the file is absent so the build still runs.
    """
    if not FINANCE_ACTUALS_PATH.exists():
        print(f"  Finance actuals file not found ({FINANCE_ACTUALS_PATH.name}); skipping finance rows.")
        return {}

    result = {}
    with open(FINANCE_ACTUALS_PATH, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            period = (row.get('PERIOD') or '').strip()
            category = (row.get('CATEGORY') or '').strip()
            channel = (row.get('CHANNEL') or '').strip()
            mdf_split = (row.get('MDF_SPLIT') or '').strip()
            amt = num(row.get('SPEND'))
            if not period or not category or not channel or amt is None:
                continue
            try:
                mon_abbr, yr = period.split()
                ym = (int(yr), MONTH_ABBR[mon_abbr[:3]])
            except (ValueError, KeyError):
                continue
            d = result.setdefault((channel, category), {})
            d[ym] = d.get(ym, 0.0) + amt
            if mdf_split and mdf_split != 'N/A - Promos':
                dsplit = result.setdefault((channel, category, mdf_split), {})
                dsplit[ym] = dsplit.get(ym, 0.0) + amt
    return result


def apply_finance_actuals(ws, finance_actuals, max_col):
    """Write monthly Finance Actuals into the dashboard 'Finance Actuals (Monthly)'
    rows. Same placement as Budgeted: LAST Mon-Sun week of each month; $ → $k.

    ACCUMULATES into each (row, col): multiple categories can map to the same row
    (Advertising + Other Mktg both → the channel's Total Marketing finance row),
    so their dollars are summed, not overwritten. FINANCE_ROW_MAP keys can be
    either 2-tuples (channel, category) or 3-tuples (channel, category,
    mdf_split) — passed straight through to finance_actuals.get(), so key
    shape doesn't matter here (2026-08-25, Amazon Non-MDF/MDF split)."""
    acc = {}  # (dash_row, col) → $ summed
    for key, dash_row in FINANCE_ROW_MAP.items():
        for (year, month), val in finance_actuals.get(key, {}).items():
            col = last_mon_of_month_col(year, month)
            if col is None or col > max_col + 30:
                continue
            acc[(dash_row, col)] = acc.get((dash_row, col), 0.0) + val
    for (dash_row, col), val in acc.items():
        ws.cell(dash_row, col).value = val / 1000.0  # $ → $k
    return len(acc)


def apply_pct_of_budget_formulas(ws, all_cols, max_col):
    """As % of budget rows — MONTHLY, one value per COMPLETE month, placed at that
    month's last Mon-Sun column (same column as Budgeted / Finance Actuals).

    Three rows per marketing block (only Promo + Total Marketing carry them):
      - this month: month finance              / month budget
      - QTD:        SUM(finance quarter-to-date) / FULL-quarter budget (all 3 months)
      - YTD:        SUM(finance year-to-date)    / FULL-year  budget (all 12 months)

    Numerator = the block's Finance Actuals (monthly) row for ALL three rows
    (user 2026-06-29: % of budget compares monthly GL finance to budget, not the
    bottom-up weekly actual). Incomplete (in-progress) months left blank (gate:
    month-end col > max_col). Full-quarter / full-year budget denominators
    reference budget cells at FUTURE month-end columns too (apply_plan_targets
    fills the whole year up to ~Dec).
    """
    # Group data week-columns into months by the week's Sunday.
    month_cols = {}
    for c in all_cols:
        sun = WEEK_BASE + timedelta(days=(c - 5) * 7 + 6)
        ym = (sun.year, sun.month)
        lo, hi = month_cols.get(ym, (c, c))
        month_cols[ym] = (min(lo, c), max(hi, c))

    n = 0
    for actual, finance, budget, pct_m, pct_q, pct_y in MARKETING_BLOCKS:
        num_row = finance   # all three % rows use the monthly finance row
        # Clear stale values across the data range first.
        for col in range(5, max_col + 1):
            ws.cell(pct_m, col).value = None
            ws.cell(pct_q, col).value = None
            ws.cell(pct_y, col).value = None
        for (y, m), (first_col, _last_col) in month_cols.items():
            mend = last_mon_of_month_col(y, m)
            if mend is None or mend > max_col:
                continue  # month not yet complete → blank
            Lm = col_letter(mend)
            Lf = col_letter(first_col)
            # quarter bounds
            q0 = ((m - 1) // 3) * 3 + 1          # first month of the quarter
            q_months = [q0, q0 + 1, q0 + 2]
            qfm_cols = month_cols.get((y, q0))
            q_start_col = qfm_cols[0] if qfm_cols else first_col
            Lqs = col_letter(q_start_col)
            q_bud = "+".join(f"{col_letter(last_mon_of_month_col(y, qm))}{budget}"
                             for qm in q_months)
            Ldec = col_letter(last_mon_of_month_col(y, 12))
            # this month
            ws.cell(pct_m, mend).value = (
                f"=IFERROR(SUM({Lf}{num_row}:{Lm}{num_row})/{Lm}{budget},\"\")")
            # QTD ÷ full-quarter budget
            ws.cell(pct_q, mend).value = (
                f"=IFERROR(SUM({Lqs}{num_row}:{Lm}{num_row})/({q_bud}),\"\")")
            # YTD ÷ full-annual budget
            ws.cell(pct_y, mend).value = (
                f"=IFERROR(SUM($E${num_row}:{Lm}{num_row})/SUM($E${budget}:{Ldec}{budget}),\"\")")
            n += 3
    return n


def clear_other_marketing_gray(ws):
    """Remove the gray-shaded fill from Other Marketing rows in the template.
    User asked to remove these once real data is flowing in.
    """
    blank_fill = PatternFill(fill_type=None)
    for r in OTHER_MARKETING_GRAY_ROWS:
        for c in range(1, 36):  # cover col A through full week range
            ws.cell(r, c).fill = blank_fill


# ============================================================
# Load CSV + apply Total Retail TTS-inclusive recomputation
# ============================================================

def load_data(csv_path):
    data = {}
    with open(csv_path, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            data[(row['WEEK_START'], row['DASHBOARD_CHANNEL'])] = row
    return data


def recompute_total_retail_including_tts(data):
    """For SUM-based fields only, override CSV's Total Retail row to include TTS US.

    Reason: SQL's total_retail rollup CTE excluded TTS until 2026-05-31. Once the SQL
    is re-run after that change, this becomes a no-op. New_customer fields are NOT
    touched (separate source — `acquisition_channel='Other'` bucket already includes TTS).
    """
    weeks = sorted({ws for (ws, _) in data.keys()})
    for w in weeks:
        tr_key = (w, 'Total Retail')
        if tr_key not in data:
            continue
        for field in SUM_FIELDS:
            total = 0
            any_val = False
            for ch in SUM_CHANNELS:
                v = num(data.get((w, ch), {}).get(field))
                if v is not None:
                    total += v
                    any_val = True
            data[tr_key][field] = str(total) if any_val else ''


# ============================================================
# Excel generator
# ============================================================

def build_excel(data, dst_path, marketing_plan, promo_plan, finance_actuals=None, snapshot_date=None):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['Dashboard']
    # Cutoff for "complete" weeks: a week is complete if its Sunday < snapshot.
    # Targets (sell-in / sell-through) are written for the FULL year; actuals,
    # single-week %, cumulative %, and NC are written only through this week.
    snap = date.fromisoformat(snapshot_date) if snapshot_date else date.today()

    # E1 = Monday of the first Mon-Sun week (header row is now "Weeks starting");
    # F1.. = =E1+7 in the template. (Was Sunday-ending before 2026-07-14.)
    ws['E1'] = WEEK_BASE
    # (Template now has correct D4='#' and 'TikTok Shop' / typos fixed by user —
    # the old D4 / A-cell overrides were removed 2026-06-29.)
    # Relabel '% of budget - this week' → '- this month' (label only; harmless if
    # the template already says 'this month').
    for blk in MARKETING_BLOCKS:
        pct_m = blk[3]
        ws.cell(pct_m, 2).value = "As % of budget - this month"

    stats = {'values': 0, 'formulas': 0, 'skipped': 0}

    def put_val(r, c, v):
        if v is not None:
            ws.cell(r, c).value = v
            stats['values'] += 1

    def put_formula(r, c, f):
        ws.cell(r, c).value = f
        stats['formulas'] += 1

    def sum_formula_5ch(col, metric_key):
        """Excel formula: =E27+E71+E100+E129+E153 (5 retail channels) per metric."""
        L = col_letter(col)
        return '=' + '+'.join(f"{L}{CH_ROWS[ch][metric_key]}" for ch in SUM_CHANNELS)

    def div_formula(col, num_row, den_row):
        L = col_letter(col)
        return f"=IFERROR({L}{num_row}/{L}{den_row},\"\")"

    # Pass 1: per-channel hardcoded values + per-channel % formulas
    for (ws_str, ch), row in data.items():
        col = week_col(ws_str)
        if col is None:
            stats['skipped'] += 1
            continue
        complete = (date.fromisoformat(ws_str) + timedelta(days=6)) < snap   # week fully in the past
        si_a = num(row['SELL_IN_GROSS_REV_ACTUAL'])
        si_t = num(row['SELL_IN_GROSS_REV_TARGET'])
        st_a = num(row['SELL_THROUGH_UNITS_ACTUAL'])
        st_t = num(row['SELL_THROUGH_UNITS_TARGET'])
        st_cam = num(row['SELL_THROUGH_UNITS_CAM_ACTUAL'])
        st_nc = num(row['SELL_THROUGH_UNITS_NONCAM_ACTUAL'])
        nc_a = num(row['NEW_CUSTOMERS_ACTUAL'])
        nc_cam = num(row['NEW_CUSTOMERS_CAM_ACTUAL'])
        nc_nc = num(row['NEW_CUSTOMERS_NONCAM_ACTUAL'])
        pr = num(row['PROMO_SPEND_ACTUAL'])
        ao = num(row['ADS_ONSITE_ACTUAL'])
        ad = num(row['ADS_DSP_ACTUAL'])
        at = num(row['ADS_TOTAL_ACTUAL'])
        k = lambda v: v / 1000 if v is not None else None

        if ch == 'All Channels':
            if complete:
                put_val(3, col, k(si_a))
                put_val(4, col, st_a)
                put_val(5, col, nc_a)
        elif ch == 'Total Retail':
            # SUM rows handled in Pass 2 as formulas; here only new_customer cells
            if complete:
                put_val(TR_ROWS['nc_a'], col, nc_a)
                put_val(TR_ROWS['nc_cam'], col, nc_cam)
                put_val(TR_ROWS['nc_nc'], col, nc_nc)
        elif ch in CH_ROWS:
            R = CH_ROWS[ch]
            # Targets — written for the FULL year (incl. future weeks).
            put_val(R['si_t'], col, k(si_t))
            if col == 5 and st_t is None:
                st_t = st_a   # first week (2025-12-29) has no sell-through target → use actual as target
            put_val(R['st_t'], col, st_t)
            # Actuals + per-week % — only for complete weeks.
            if complete:
                put_val(R['si_a'], col, k(si_a))
                put_formula(R['si_pct'], col, div_formula(col, R['si_a'], R['si_t']))
                put_val(R['st_a'], col, st_a)
                put_formula(R['st_pct'], col, div_formula(col, R['st_a'], R['st_t']))
                put_val(R['st_cam'], col, st_cam)
                put_val(R['st_nc'], col, st_nc)
                if 'nc_a' in R:
                    put_val(R['nc_a'], col, nc_a)   # Amazon NTB actual (R43); ratio rows stay blank
                if 'promo' in R and k(pr) is not None:
                    put_val(R['promo'], col, k(pr))
                if ch == 'Amazon':
                    put_val(R['ads_on'], col, k(ao))
                    put_val(R['ads_dsp'], col, k(ad))
                    L = col_letter(col)
                    put_formula(R['ads_tot'], col, f"={L}{R['ads_on']}+{L}{R['ads_dsp']}")
                elif ch in ('TikTok Shop', 'Home Depot', 'Walmart'):
                    put_val(R['ads_tot'], col, k(at))

    # Determine dynamic column range from CSV data — week count varies each refresh
    # as more weeks become complete (yesterday's Sun joins the eligible set tomorrow).
    all_cols = sorted({week_col(w) for (w, _) in data.keys() if week_col(w) is not None})
    if not all_cols:
        wb.save(dst_path)
        return stats
    max_col = max(all_cols)  # inclusive (full year incl. future target weeks)
    # Last complete week column (Sunday < snapshot). Targets go to max_col; actuals
    # / single-week % / cumulative % / NC only through last_actual_col.
    complete_cols = [c for c in all_cols if (WEEK_BASE + timedelta(days=(c - 5) * 7 + 6)) < snap]
    last_actual_col = max(complete_cols) if complete_cols else max_col
    col_range = range(5, last_actual_col + 1)   # complete weeks only

    # Pass 2: Total Retail SUM rows. Targets (si_t / st_t) for the FULL year;
    # actuals / %s only through the last complete week.
    for col in all_cols:
        put_formula(TR_ROWS['si_t'], col, sum_formula_5ch(col, 'si_t'))
        put_formula(TR_ROWS['st_t'], col, sum_formula_5ch(col, 'st_t'))
    for col in col_range:
        put_formula(TR_ROWS['si_a'], col, sum_formula_5ch(col, 'si_a'))
        put_formula(TR_ROWS['si_pct'], col, div_formula(col, TR_ROWS['si_a'], TR_ROWS['si_t']))
        put_formula(TR_ROWS['st_a'], col, sum_formula_5ch(col, 'st_a'))
        put_formula(TR_ROWS['st_pct'], col, div_formula(col, TR_ROWS['st_a'], TR_ROWS['st_t']))
        put_formula(TR_ROWS['st_cam'], col, sum_formula_5ch(col, 'st_cam'))
        put_formula(TR_ROWS['st_nc'], col, sum_formula_5ch(col, 'st_nc'))

    # Pass 2b: New Customers. Flat 52% target on TR (R23) + Amazon (R47). Total
    # Retail also gets "% new customers - YTD" (R21 = ΣNC/Σunits cumulative) and
    # "% of Target - YTD" (R22 = R21 − R23, pp). Amazon ratio rows R44/45/46 stay
    # blank (NTB orders vs sell-through units = mixed denominator).
    T = TR_ROWS
    for col in col_range:
        L = col_letter(col)
        for t_row in NC_TARGET_ROWS:                 # flat 52% target, every complete week
            ws.cell(t_row, col).value = NC_TARGET_PCT
            stats['values'] += 1
        put_formula(T['nc_occ_ytd'], col,
            f"=IFERROR(SUM($E${T['nc_a']}:{L}{T['nc_a']})/SUM($E${T['st_a']}:{L}{T['st_a']}),\"\")")
        put_formula(T['nc_pct'], col, f"=IFERROR({L}{T['nc_occ_ytd']}-{L}{T['nc_t']},\"\")")

    # Pass 3: % of Total — YTD formulas (channel YTD share of All Channels)
    for target_row, num_row, den_row in YTD_PCT_ROWS:
        for col in col_range:
            L = col_letter(col)
            put_formula(target_row, col,
                f"=IFERROR(SUM($E${num_row}:{L}{num_row})/SUM($E${den_row}:{L}{den_row}),\"\")")

    # Pass 4: % of Target — YTD = cumulative actual ÷ cumulative target, INCLUDING
    # the first week (2026-06-30): sell-through's first week (2025-12-29) had no
    # target, now backfilled with that week's actual (see Pass 1), so a plain
    # SUM/SUM over all weeks is correct (no week left with actual-but-blank-target).
    for pct_ytd_row, actual_row, target_row in TARGET_YTD_PCT_ROWS:
        for col in col_range:
            L = col_letter(col)
            put_formula(pct_ytd_row, col,
                f"=IFERROR(SUM($E${actual_row}:{L}{actual_row})"
                f"/SUM($E${target_row}:{L}{target_row}),\"\")")

    # Plan targets — Promo + Marketing budgets (monthly grain → placed at
    # LAST Mon-Sun week of each month, into the budgeted_monthly row of each block)
    plan_max_col = max(max_col, last_mon_of_month_col(2026, 12) or max_col)
    plan_written = apply_plan_targets(ws, marketing_plan, promo_plan, max_col)
    stats['plan_cells'] = plan_written

    # Finance Actuals (Monthly) — real GL spend from Wyze_Marketing_Cost_by_Channel.
    # Same monthly→LAST-Mon-Sun-week placement as Budgeted, so the % of budget
    # formulas (finance ÷ budget, same column) line up. Only TikTok Shop today.
    fin_written = apply_finance_actuals(ws, finance_actuals or {}, max_col)
    stats['finance_cells'] = fin_written

    # % of budget — this month / YTD: monthly weekly-actual ÷ monthly Budgeted,
    # one value per COMPLETE month at the month-end column (incomplete months blank).
    # Gate by last_actual_col (complete weeks) so % of budget only fills complete
    # months — not the future months that now exist (full-year target weeks).
    pct_written = apply_pct_of_budget_formulas(ws, all_cols, last_actual_col)
    stats['formulas'] += pct_written

    # Remove gray fill from Other Marketing rows (safety net — user may have
    # already de-grayed in template).
    clear_other_marketing_gray(ws)

    # Apply number formatting (integer everywhere; reads col D Unit).
    # Apply to col E through max possible plan column (~ col AI for Dec) to
    # cover plan months beyond visible data weeks.
    apply_number_formats(ws, plan_max_col)

    # Unify each data row's number cells with its row-name cell's style (color /
    # bold / italic) so a row reads as one visual band — easier to scan across
    # the dense week grid (user 2026-06-22). Number format is left untouched.
    apply_row_label_styles(ws, plan_max_col)

    # Cell comment on each sell-through "Target" row label, explaining the
    # first-week backfill (no forecast for wk ending 2026-01-04 → use that week's
    # actual as the target). Hover-only; doesn't affect layout.
    from openpyxl.comments import Comment
    st_target_note = Comment(
        "Sell-through has no forecast target for the first week (week ending "
        "2026-01-04, i.e. 2025-12-29). For every channel that week's Target is set "
        "equal to that week's Actual, so its '% of Target' = 100% and the YTD "
        "'% of Target' includes the first week.",
        "CAC tracker")
    st_target_note.width = 320
    st_target_note.height = 140
    # (2026-08-25: this list previously pointed at each channel's Cam-breakdown
    # row, one below the real Target row, for all 4 non-Amazon/non-TR channels
    # — fixed here while remapping rows for the Non-MDF/MDF insertion below.)
    for r in (16, 39, 101, 137, 173, 203):   # sell-through Target row labels
        ws.cell(r, 2).comment = Comment(st_target_note.text, st_target_note.author,
                                        height=140, width=320)

    wb.save(dst_path)
    return stats


def apply_row_label_styles(ws, max_col):
    """Make each data row's number cells match its label cell, so the whole row
    reads as one band. Copies font color / bold / italic always, and the label's
    FILL when it has one (so a row the user greys out — e.g. the Amazon NC block —
    greys across the data cells too). The label cell is col A if it has text, else
    col B. Section-header rows (no Unit in col D) are skipped. Data cells keep their
    own font name/size and number_format."""
    from copy import copy
    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 4).value is None:   # no Unit → header / spacer row
            continue
        lab = ws.cell(row, 1) if ws.cell(row, 1).value not in (None, '') else ws.cell(row, 2)
        lf = lab.font
        lab_has_fill = lab.fill is not None and lab.fill.patternType is not None
        for col in range(5, max_col + 1):
            cell = ws.cell(row, col)
            cf = cell.font
            cell.font = Font(name=cf.name, size=cf.size,
                             bold=lf.bold, italic=lf.italic,
                             color=copy(lf.color) if lf.color else None)
            if lab_has_fill:                # carry the label's fill (e.g. grey) across
                cell.fill = copy(lab.fill)


def apply_number_formats(ws, max_col):
    """Set Excel display formats per row, based on Unit column (D).

    All numbers are INTEGER (no decimals) per user 2026-06-18:
      - Unit = '%' (any row)            → '0%'    (integer percent)
      - Unit in ('$k','#','# ppl')      → '#,##0' (integer with thousand separator)

    Section-header rows (no Unit in col D) are skipped. Format applied from
    col E through `max_col` (the last week column present in this CSV).
    """
    INT_FMT = '#,##0'
    PCT_FMT = '0%'

    # Cover every data row through the full row extent (template is 212 rows).
    for row in range(3, ws.max_row + 1):
        unit = ws.cell(row, 4).value
        if unit is None:
            continue
        unit_str = str(unit).strip()
        if unit_str in ('%', 'pp'):       # pp shown with a % sign too (e.g. "5%")
            fmt = PCT_FMT
        elif unit_str in ('$k', '#', '# ppl'):
            fmt = INT_FMT
        else:
            continue
        for col in range(5, max_col + 1):
            ws.cell(row, col).number_format = fmt


# ============================================================
# HTML generator
# ============================================================

def build_html(data, dst_path, snapshot_date, marketing_plan, promo_plan, finance_actuals=None):
    finance_actuals = finance_actuals or {}
    today = date.fromisoformat(snapshot_date)
    weeks = sorted({
        ws for (ws, _) in data.keys()
        if ws >= WEEK_BASE.isoformat()
        and (date.fromisoformat(ws) + timedelta(days=6)) < today
    })
    if not weeks:
        raise SystemExit("No weeks to render — check CSV contents or snapshot date.")

    week_labels = [
        (w, date.fromisoformat(w).strftime('%b %-d'))   # Monday (week starting)
        for w in weeks
    ]
    most_recent_ws = max(weeks)
    most_recent_label = (
        f"{date.fromisoformat(most_recent_ws).strftime('%b %-d')} – "
        f"{(date.fromisoformat(most_recent_ws) + timedelta(days=6)).strftime('%b %-d, %Y')}"
    )

    # Cumulative YTD % calculator per channel × field, week-by-week
    def cum_pct(ch, field, denom_ch='Total Retail'):
        out = []
        num_cum = 0
        den_cum = 0
        for w in weeks:
            nv = num(data.get((w, ch), {}).get(field)) or 0
            dv = num(data.get((w, denom_ch), {}).get(field)) or 0
            num_cum += nv
            den_cum += dv
            out.append(num_cum / den_cum if den_cum else None)
        return out

    # Cumulative "% of Target — YTD": same channel, actual ÷ target, week-by-week.
    # Weeks with a MISSING target (e.g. sell-through target blank for the first
    # week 2025-12-29) are excluded from BOTH sides — including their actual would
    # inflate the ratio. Mirrors the Excel SUMIF(target<>"", actual) logic.
    def cum_pct_target(ch, actual_field, target_field):
        out = []
        a_cum = 0.0
        t_cum = 0.0
        for w in weeks:
            tv = num(data.get((w, ch), {}).get(target_field))
            if tv is not None:
                a_cum += num(data.get((w, ch), {}).get(actual_field)) or 0
                t_cum += tv
            out.append(a_cum / t_cum if t_cum else None)
        return out

    def fmt_k(v):
        return '' if v is None else f"${v/1000:,.0f}k"

    def fmt_num(v):
        return '' if v is None else f"{v:,.0f}"

    def fmt_pct(v, dec=0):
        return '' if v is None else f"{v*100:.{dec}f}%"

    def pct_class(v):
        if v is None:
            return ''
        return 'pos' if v >= 1.0 else ('neg' if v < 0.95 else 'mid')

    def get(ws_str, ch, field):
        r = data.get((ws_str, ch))
        if not r:
            return None
        return num(r.get(field))

    # Build lookup: (year, month) → week_start_iso of last Mon-Sun week ending in that month
    last_mon_per_month = {}
    for w in weeks:
        d = date.fromisoformat(w)
        sun = d + timedelta(days=6)
        ym = (sun.year, sun.month)
        if ym not in last_mon_per_month or d > date.fromisoformat(last_mon_per_month[ym]):
            last_mon_per_month[ym] = w

    # Map (HTML section channel, HTML metric label) → plan source spec
    # spec = ('promo', plan_channel_name, None)
    #     or ('marketing', plan_channel_name, [bucket1, bucket2, ...])
    #     or ('marketing_multi', [channel1, channel2, ...], None)
    PLAN_MAP = {
        ('Amazon',      'Promo'):                  ('promo',     'Amazon 1P', None),
        ('Amazon',      'Total Marketing'):        ('marketing_multi', ['Amazon 1P'], None),
        # 2026-08-25: Non-MDF / MDF — a second, independent breakdown of the
        # same Total Marketing total (the other being Ads + Other Marketing
        # below). Non-MDF = Total Marketing's 7 buckets other than MDF.
        ('Amazon',      'Non-MDF'):                 ('marketing', 'Amazon 1P', ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads', 'DSP', 'Influencer', 'Strategic Vendor Service - FTE', 'Prime Video Ads', 'Affiliate']),
        ('Amazon',      'MDF'):                     ('marketing', 'Amazon 1P', ['MDF']),
        ('Walmart',     'Total Marketing'):        ('marketing_multi', ['Walmart'], None),
        ('Home Depot',  'Total Marketing'):        ('marketing_multi', ['Home Depot'], None),
        ('TikTok Shop', 'Total Marketing'):        ('marketing_multi', ['TikTok Shop'], None),
        ('Amazon',      'Ads — On-site'):          ('marketing', 'Amazon 1P', ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads']),
        ('Amazon',      'Ads — DSP'):              ('marketing', 'Amazon 1P', ['DSP']),
        ('Amazon',      'Ads (On-site + DSP)'):    ('marketing', 'Amazon 1P', ['Non-Brand - Sponsored Ads', 'Brand - Sponsored Ads', 'DSP']),
        ('Amazon',      'Other Marketing'):        ('marketing', 'Amazon 1P', ['MDF', 'Influencer', 'Strategic Vendor Service - FTE', 'Prime Video Ads', 'Affiliate']),
        ('Walmart',     'Promo'):                  ('promo',     'Walmart',   None),
        ('Walmart',     'Ads'):                    ('marketing', 'Walmart',   ['Ad Spend']),
        ('Walmart',     'Other Marketing'):        ('marketing', 'Walmart',   ['In-Store Display', 'Content']),
        ('Home Depot',  'Promo'):                  ('promo',     'Home Depot', None),
        ('Home Depot',  'Ads'):                    ('marketing', 'Home Depot', ['Marketing package']),
        ('Home Depot',  'Other Marketing'):        ('marketing', 'Home Depot', ['In-Store Display', 'Events', 'Brand Advocate']),
        ('Other Omni',  'Promo'):                  ('promo',     'Best Buy',  None),
        ('Other Omni',  'Total Marketing'):        ('marketing_multi', ['Best Buy', 'Costco', 'Other Retail'], None),
        ('TikTok Shop', 'Promo'):                  ('promo',     'TikTok Shop', None),
        ('TikTok Shop', 'Ads (GMV Max)'):          ('marketing', 'TikTok Shop', ['Ad Spend']),
        ('TikTok Shop', 'Other Marketing'):        ('marketing', 'TikTok Shop', ['Affiliate Commisions', 'Free Samples', 'Livestreaming', 'Affiliate Maintenance and Tools', 'Video']),
    }

    def budgeted_monthly_for(ch, metric_label):
        """Return list aligned with `weeks`, monthly $k at last-Mon-of-month col, else ''."""
        spec = PLAN_MAP.get((ch, metric_label))
        if not spec:
            return ['' for _ in weeks]
        kind, ch_or_list, buckets = spec
        monthly = {}
        if kind == 'promo':
            monthly = dict(promo_plan.get(ch_or_list, {}))
        elif kind == 'marketing':
            for b in buckets:
                for ym, v in marketing_plan.get((ch_or_list, b), {}).items():
                    monthly[ym] = monthly.get(ym, 0) + v
        elif kind == 'marketing_multi':
            for ch_name in ch_or_list:
                for (c, b), m in marketing_plan.items():
                    if c != ch_name:
                        continue
                    for ym, v in m.items():
                        monthly[ym] = monthly.get(ym, 0) + v
        out = []
        for w in weeks:
            d = date.fromisoformat(w)
            sun = d + timedelta(days=6)
            ym = (sun.year, sun.month)
            target_w = last_mon_per_month.get(ym)
            if target_w == w and ym in monthly:
                # plan dict is in $; fmt_k expects $ and renders as $k
                out.append(fmt_k(monthly[ym]))
            else:
                out.append('')
        return out

    def finance_monthly_for(ch, metric_label):
        """Finance Actuals (Monthly) aligned with `weeks`: monthly $ at the
        last-Mon-of-month col, else ''. Only mapped (ch, metric) have data."""
        key = FINANCE_HTML_MAP.get((ch, metric_label))
        monthly = finance_actuals.get(key, {}) if key else {}
        out = []
        for w in weeks:
            sun = date.fromisoformat(w) + timedelta(days=6)
            ym = (sun.year, sun.month)
            if last_mon_per_month.get(ym) == w and ym in monthly:
                out.append(fmt_k(monthly[ym]))   # $ → $k
            else:
                out.append('')
        return out

    # dashboard channel → finance source channel (for Total Marketing finance sum)
    FINANCE_CH = {'Amazon': 'Amazon 1P', 'Walmart': 'Walmart', 'Home Depot': 'Home Depot',
                  'Other Omni': 'Other Retail', 'TikTok Shop': 'TTS'}

    def finance_total_mktg(ch):
        """Total Marketing finance = Advertising + Other Mktg (GL) summed, placed
        at the last-Mon-of-month column. Mirrors the Excel Total Marketing row."""
        fch = FINANCE_CH.get(ch)
        monthly = {}
        for cat in ('Advertising', 'Other Mktg'):
            for ym, v in (finance_actuals.get((fch, cat), {}) if fch else {}).items():
                monthly[ym] = monthly.get(ym, 0) + v
        out = []
        for w in weeks:
            sun = date.fromisoformat(w) + timedelta(days=6)
            ym = (sun.year, sun.month)
            out.append(fmt_k(monthly[ym]) if (last_mon_per_month.get(ym) == w and ym in monthly) else '')
        return out

    def finance_mdf_split(ch, split):
        """Non-MDF or MDF finance (2026-08-25) — sums the 3-tuple granular keys
        load_finance_actuals() writes alongside (channel, category). Amazon
        only; other channels have no MDF_SPLIT data so this returns all ''."""
        fch = FINANCE_CH.get(ch)
        monthly = {}
        for cat in ('Advertising', 'Other Mktg'):
            for ym, v in (finance_actuals.get((fch, cat, split), {}) if fch else {}).items():
                monthly[ym] = monthly.get(ym, 0) + v
        out = []
        for w in weeks:
            sun = date.fromisoformat(w) + timedelta(days=6)
            ym = (sun.year, sun.month)
            out.append(fmt_k(monthly[ym]) if (last_mon_per_month.get(ym) == w and ym in monthly) else '')
        return out

    rows = []

    def add_row(label, unit, values, kind='', indent=0):
        rows.append({'label': label, 'unit': unit, 'values': values, 'kind': kind, 'indent': indent})

    def add_section_header(label):
        rows.append({'section_header': label})

    def add_finance_row(ch, metric_label):
        """Add 'Finance Actuals (Monthly)' row (real GL spend) if finance data
        exists for this (channel, metric). Only TikTok Shop is mapped today."""
        vals = finance_monthly_for(ch, metric_label)
        if any(v != '' for v in vals):
            add_row("Finance Actuals (Monthly)", "$k", vals, kind='finance', indent=1)

    def add_budget_row(ch, metric_label):
        """Add Finance Actuals (Monthly) then Budgeted (Monthly) right after the
        actual row, so each metric block reads actual → finance → budget."""
        add_finance_row(ch, metric_label)
        vals = budgeted_monthly_for(ch, metric_label)
        if any(v != '' for v in vals):
            add_row("Budgeted (Monthly)", "$k", vals, kind='budget', indent=1)

    def add_budget_only(ch, metric_label):
        """Budgeted (Monthly) row with NO finance row above it — used for the
        Ads / Other Marketing detail blocks (finance lives in Total Marketing)."""
        vals = budgeted_monthly_for(ch, metric_label)
        if any(v != '' for v in vals):
            add_row("Budgeted (Monthly)", "$k", vals, kind='budget', indent=1)

    add_section_header("All Channels")
    add_row("Sell-in value", "$k",
        [fmt_k(get(w, 'All Channels', 'SELL_IN_GROSS_REV_ACTUAL')) for w in weeks])
    add_row("Sell-through units", "#",
        [fmt_num(get(w, 'All Channels', 'SELL_THROUGH_UNITS_ACTUAL')) for w in weeks])
    add_row("New Customers", "# ppl",
        [fmt_num(get(w, 'All Channels', 'NEW_CUSTOMERS_ACTUAL')) for w in weeks])

    def build_channel_section(ch_label, ch, has_new_customers=False, has_amazon_ads=False, has_tts_ads=False, has_hd_ads=False, show_ytd_pct=False):
        add_section_header(ch_label)
        # Sell-in
        add_row("Sell-in value", "$k",
            [fmt_k(get(w, ch, 'SELL_IN_GROSS_REV_ACTUAL')) for w in weeks])
        if show_ytd_pct:
            add_row("% of Total Retail — YTD", "%",
                [fmt_pct(v) for v in cum_pct(ch, 'SELL_IN_GROSS_REV_ACTUAL')],
                kind='pct-ytd', indent=1)
        add_row("Target", "$k",
            [fmt_k(get(w, ch, 'SELL_IN_GROSS_REV_TARGET')) for w in weeks],
            kind='target', indent=1)
        pcts = [safediv(get(w, ch, 'SELL_IN_GROSS_REV_ACTUAL'),
                        get(w, ch, 'SELL_IN_GROSS_REV_TARGET')) for w in weeks]
        add_row("% of Target", "%",
            [(fmt_pct(v), pct_class(v)) for v in pcts], kind='pct', indent=1)
        add_row("% of Target — YTD", "%",
            [(fmt_pct(v), pct_class(v)) for v in
             cum_pct_target(ch, 'SELL_IN_GROSS_REV_ACTUAL', 'SELL_IN_GROSS_REV_TARGET')],
            kind='pct-ytd', indent=1)
        # Sell-through
        add_row("Sell-through units", "#",
            [fmt_num(get(w, ch, 'SELL_THROUGH_UNITS_ACTUAL')) for w in weeks])
        if show_ytd_pct:
            add_row("% of Total Retail — YTD", "%",
                [fmt_pct(v) for v in cum_pct(ch, 'SELL_THROUGH_UNITS_ACTUAL')],
                kind='pct-ytd', indent=1)
        add_row("Target", "#",
            [fmt_num(get(w, ch, 'SELL_THROUGH_UNITS_TARGET')) for w in weeks],
            kind='target', indent=1)
        pcts = [safediv(get(w, ch, 'SELL_THROUGH_UNITS_ACTUAL'),
                        get(w, ch, 'SELL_THROUGH_UNITS_TARGET')) for w in weeks]
        add_row("% of Target", "%",
            [(fmt_pct(v), pct_class(v)) for v in pcts], kind='pct', indent=1)
        add_row("% of Target — YTD", "%",
            [(fmt_pct(v), pct_class(v)) for v in
             cum_pct_target(ch, 'SELL_THROUGH_UNITS_ACTUAL', 'SELL_THROUGH_UNITS_TARGET')],
            kind='pct-ytd', indent=1)
        add_row("Cam", "#",
            [fmt_num(get(w, ch, 'SELL_THROUGH_UNITS_CAM_ACTUAL')) for w in weeks], indent=2)
        add_row("Non Cam", "#",
            [fmt_num(get(w, ch, 'SELL_THROUGH_UNITS_NONCAM_ACTUAL')) for w in weeks], indent=2)
        # New customers
        if has_new_customers:
            nc_label = "New Customers (Ads attributed)" if ch == 'Amazon' else "New Customers"
            add_row(nc_label, "# ppl",
                [fmt_num(get(w, ch, 'NEW_CUSTOMERS_ACTUAL')) for w in weeks])
            if ch == 'Total Retail':
                add_row("Cam", "# ppl",
                    [fmt_num(get(w, ch, 'NEW_CUSTOMERS_CAM_ACTUAL')) for w in weeks], indent=2)
                add_row("Non Cam", "# ppl",
                    [fmt_num(get(w, ch, 'NEW_CUSTOMERS_NONCAM_ACTUAL')) for w in weeks], indent=2)
            # NC target = flat 52%; "% of Target" rows are percentage-points (pp)
            # = actual share − 52%. Total Retail gets the full set; Amazon shows
            # actual + target only (ratio rows blank — NTB orders vs units).
            if ch == 'Total Retail':
                # % new customers - YTD (cumulative NC ÷ cumulative units) + Target
                # (flat 52%) + % of Target - YTD (pp = cumulative occ − 52%). No
                # single-week rows (binding lag makes weekly occupancy misleading).
                occ_ytd, tgt_ytd = [], []
                na = ta = 0.0
                for w in weeks:
                    na += get(w, ch, 'NEW_CUSTOMERS_ACTUAL') or 0
                    ta += get(w, ch, 'SELL_THROUGH_UNITS_ACTUAL') or 0
                    occ_ytd.append(fmt_pct(na / ta) if ta else '')
                    tgt_ytd.append(fmt_pct(na / ta - NC_TARGET_PCT) if ta else '')
                add_row("% new customers - YTD", "%", occ_ytd, kind='pct-ytd', indent=1)
                add_row("Target", "%", [fmt_pct(NC_TARGET_PCT) for _ in weeks], kind='target', indent=1)
                add_row("% of Target - YTD", "pp", tgt_ytd, kind='pct-ytd', indent=1)
            elif ch == 'Amazon':
                add_row("Target", "%", [fmt_pct(NC_TARGET_PCT) for _ in weeks], kind='target', indent=1)
        # Promo (weekly actual + finance + budget)
        add_row("Promo", "$k",
            [fmt_k(get(w, ch, 'PROMO_SPEND_ACTUAL')) for w in weeks])
        add_budget_row(ch, 'Promo')

        # Total Marketing — no weekly feed. Finance = Advertising + Other Mktg (GL)
        # summed; budget = all ads + other-marketing plan buckets. Ads/Other shown
        # as indented detail below (budget only — finance lives in this row).
        blank_row = ['' for _ in weeks]

        def add_label_only(label, indent=1):
            add_row(label, "$k", blank_row, indent=indent)

        add_row("Total Marketing", "$k", blank_row)
        tm_fin = finance_total_mktg(ch)
        if any(v != '' for v in tm_fin):
            add_row("Finance Actuals (Monthly)", "$k", tm_fin, kind='finance', indent=1)
        add_budget_only(ch, 'Total Marketing')

        # Non-MDF / MDF (2026-08-25, Amazon only) — a SECOND, independent
        # breakdown of the same Total Marketing total (the other being the
        # Ads + Other Marketing detail below). Not additive with it — both
        # cuts describe the same dollars from a different angle.
        if ch == 'Amazon':
            add_row("Non-MDF", "$k", blank_row, indent=1)
            nm_fin = finance_mdf_split(ch, 'Non-MDF')
            if any(v != '' for v in nm_fin):
                add_row("Finance Actuals (Monthly)", "$k", nm_fin, kind='finance', indent=1)
            add_budget_only(ch, 'Non-MDF')
            add_row("MDF", "$k", blank_row, indent=1)
            mdf_fin = finance_mdf_split(ch, 'MDF')
            if any(v != '' for v in mdf_fin):
                add_row("Finance Actuals (Monthly)", "$k", mdf_fin, kind='finance', indent=1)
            add_budget_only(ch, 'MDF')

        if has_amazon_ads:
            add_row("Ads — On-site", "$k",
                [fmt_k(get(w, ch, 'ADS_ONSITE_ACTUAL')) for w in weeks], indent=1)
            add_budget_only(ch, 'Ads — On-site')
            add_row("Ads — DSP", "$k",
                [fmt_k(get(w, ch, 'ADS_DSP_ACTUAL')) for w in weeks], indent=1)
            add_budget_only(ch, 'Ads — DSP')
            add_row("Ads (On-site + DSP)", "$k",
                [fmt_k(get(w, ch, 'ADS_TOTAL_ACTUAL')) for w in weeks], indent=1)
            add_budget_only(ch, 'Ads (On-site + DSP)')
            add_label_only("Other Marketing")
            add_budget_only(ch, 'Other Marketing')
        elif has_tts_ads:
            add_row("Ads (GMV Max)", "$k",
                [fmt_k(get(w, ch, 'ADS_TOTAL_ACTUAL')) for w in weeks], indent=1)
            add_budget_only(ch, 'Ads (GMV Max)')
            add_label_only("Other Marketing")
            add_budget_only(ch, 'Other Marketing')
        elif has_hd_ads:
            add_row("Ads", "$k",
                [fmt_k(get(w, ch, 'ADS_TOTAL_ACTUAL')) for w in weeks], indent=1)
            add_budget_only(ch, 'Ads')
            add_label_only("Other Marketing")
            add_budget_only(ch, 'Other Marketing')
        elif ch == 'Walmart':
            # Walmart Connect ads now have a weekly feed (1P+3P combined).
            add_row("Ads", "$k",
                [fmt_k(get(w, ch, 'ADS_TOTAL_ACTUAL')) for w in weeks], indent=1)
            add_budget_only(ch, 'Ads')
            add_label_only("Other Marketing")
            add_budget_only(ch, 'Other Marketing')
        # Other Omni: Total Marketing only — no Ads/Other detail.

    build_channel_section("Total Retail (Amazon 1P + Walmart + HD + Other Omni + TikTok)", "Total Retail", has_new_customers=True)
    build_channel_section("Amazon (1P only)", "Amazon", has_new_customers=True, has_amazon_ads=True, show_ytd_pct=True)
    build_channel_section("Walmart (1P+3P)", "Walmart", show_ytd_pct=True)
    build_channel_section("Home Depot (US + CA)", "Home Depot", has_hd_ads=True, show_ytd_pct=True)
    build_channel_section("Other Omni-channel", "Other Omni", show_ytd_pct=True)
    build_channel_section("TikTok Shop (US + MX)", "TikTok Shop", has_tts_ads=True, show_ytd_pct=True)

    # Render HTML
    n_weeks = len(weeks)
    parts = []
    parts.append(f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<title>Retail Business Dashboard — {snapshot_date}</title>
<style>
  :root {{
    --bg: #f6f8fa; --card: #ffffff; --ink: #1f2328; --muted: #6e7781;
    --line: #e5e7eb; --rule: #d0d7de; --section: #1f2328; --section-bg: #eef1f4;
    --target: #6e6291; --target-bg: #f5f0ff; --pos: #1a7f37; --neg: #cf222e; --mid: #9a6700;
    --ytd: #1e6091; --ytd-bg: #eef5fa; --recent: #fff8e6;
  }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Helvetica, Arial, sans-serif;
         background: var(--bg); color: var(--ink); margin: 0; padding: 24px; }}
  h1 {{ font-size: 18px; font-weight: 600; margin: 0 0 4px 0; }}
  .sub {{ color: var(--muted); font-size: 12px; margin: 0 0 16px 0; }}
  .wrap {{ background: var(--card); border: 1px solid var(--line); border-radius: 10px;
          box-shadow: 0 1px 3px rgba(0,0,0,0.04); overflow: auto; max-height: calc(100vh - 100px); }}
  table {{ border-collapse: separate; border-spacing: 0; font-size: 11.5px; min-width: 100%; }}
  thead th {{ position: sticky; top: 0; background: #fafbfc; z-index: 3;
              border-bottom: 1px solid var(--rule); padding: 8px 10px; text-align: right;
              font-weight: 600; white-space: nowrap; }}
  thead th.metric-col {{ left: 0; z-index: 4; text-align: left; min-width: 240px;
                         border-right: 1px solid var(--rule); }}
  thead th.unit-col {{ left: 240px; z-index: 4; min-width: 50px; text-align: center;
                        border-right: 1px solid var(--rule); background: #fafbfc; }}
  thead th .ws {{ display: block; font-size: 9.5px; font-weight: 400; color: var(--muted); margin-top: 2px; }}
  tbody td {{ padding: 6px 10px; text-align: right; white-space: nowrap;
              border-bottom: 1px solid var(--line); font-variant-numeric: tabular-nums; }}
  tbody td.metric-col {{ position: sticky; left: 0; background: var(--card); text-align: left;
                          font-weight: 500; border-right: 1px solid var(--rule); z-index: 2; }}
  tbody td.unit-col {{ position: sticky; left: 240px; background: var(--card);
                       text-align: center; color: var(--muted); font-size: 10.5px;
                       border-right: 1px solid var(--rule); z-index: 2; }}
  tr.section-header td {{ background: var(--section-bg); font-weight: 700; color: var(--section);
                          padding: 8px 10px; border-top: 2px solid var(--rule); }}
  tr.section-header td.metric-col {{ background: var(--section-bg); }}
  tr.target td {{ color: var(--target); background: var(--target-bg); font-size: 11px; }}
  tr.target td.metric-col, tr.target td.unit-col {{ background: var(--target-bg); }}
  tr.pct td {{ font-size: 11px; }}
  tr.pct-ytd td {{ color: var(--ytd); background: var(--ytd-bg); font-size: 11px; font-weight: 500; }}
  tr.pct-ytd td.metric-col, tr.pct-ytd td.unit-col {{ background: var(--ytd-bg); }}
  tr.budget td {{ color: var(--target); background: var(--target-bg); font-size: 11px; }}
  tr.budget td.metric-col, tr.budget td.unit-col {{ background: var(--target-bg); }}
  tr.finance td {{ color: #0f6b53; background: #ecf7f2; font-size: 11px; font-weight: 500; }}
  tr.finance td.metric-col, tr.finance td.unit-col {{ background: #ecf7f2; }}
  .pos {{ color: var(--pos); font-weight: 600; }}
  .neg {{ color: var(--neg); font-weight: 600; }}
  .mid {{ color: var(--mid); font-weight: 600; }}
  .indent-1 {{ padding-left: 24px !important; }}
  .indent-2 {{ padding-left: 40px !important; color: var(--muted); font-weight: 400; }}
  td.recent {{ background: var(--recent); }}
  td.recent.metric-col, td.recent.unit-col {{ background: var(--recent); }}
  th.recent {{ background: var(--recent) !important; }}
  .note {{ margin-top: 12px; padding: 10px 14px; background: #fffbea; border: 1px solid #d4a017;
           border-radius: 8px; font-size: 11.5px; color: #5a4000; line-height: 1.6; }}
</style></head>
<body>
<h1>Retail Business Dashboard</h1>
<p class="sub">Snapshot {snapshot_date} · Most recent complete week: {most_recent_label} · Mon-Sun weeks (header = Monday, week starting) · {n_weeks} weeks shown</p>
<div class="wrap">
<table>
<thead><tr>
  <th class="metric-col">Metric</th>
  <th class="unit-col">Unit</th>""")

    for i, (w, lbl) in enumerate(week_labels):
        cls = ' class="recent"' if i == n_weeks - 1 else ''
        parts.append(f'  <th{cls}>{lbl}<span class="ws">wk of {w}</span></th>\n')
    parts.append('</tr></thead><tbody>\n')

    for r in rows:
        if 'section_header' in r:
            parts.append(
                f'<tr class="section-header"><td class="metric-col" colspan="{2 + n_weeks}">'
                f'{r["section_header"]}</td></tr>\n')
            continue
        indent_cls = f' class="indent-{r["indent"]}"' if r['indent'] else ''
        tr_cls = ''
        if r['kind'] == 'target':
            tr_cls = ' class="target"'
        elif r['kind'] == 'pct':
            tr_cls = ' class="pct"'
        elif r['kind'] == 'pct-ytd':
            tr_cls = ' class="pct-ytd"'
        elif r['kind'] == 'budget':
            tr_cls = ' class="budget"'
        elif r['kind'] == 'finance':
            tr_cls = ' class="finance"'
        parts.append(f'<tr{tr_cls}>')
        parts.append(f'<td class="metric-col"><span{indent_cls}>{r["label"]}</span></td>')
        parts.append(f'<td class="unit-col">{r["unit"]}</td>')
        for i, v in enumerate(r['values']):
            recent_cls = ' recent' if i == n_weeks - 1 else ''
            if isinstance(v, tuple):
                val, cls = v
                cls_attr = (cls + recent_cls).strip()
                cls_attr = f' class="{cls_attr}"' if cls_attr else ''
                parts.append(f'<td{cls_attr}>{val}</td>')
            else:
                cls_attr = f' class="{recent_cls.strip()}"' if recent_cls else ''
                parts.append(f'<td{cls_attr}>{v}</td>')
        parts.append('</tr>\n')

    parts.append("""</tbody></table></div>
<div class="note">
  <strong>Notes:</strong>
  • Column header = Monday starting the Mon-Sun week. Highlighted column = most recent complete week.
  • <span class="pos">Green</span> / <span class="neg">red</span> on "% of Target" rows: ≥100% favorable; &lt;95% unfavorable.
  • <span style="background:#eef5fa;color:#1e6091;font-weight:500;padding:1px 4px;">% of Total Retail — YTD</span> rows show channel's YTD cumulative share of <b>All Channels</b> (denominator changed from Total Retail to All Channels per user 2026-05-31).
  • <b>Amazon section = 1P only</b>. Amazon 3P sell-in/sell-through exists in CSV as separate row but is excluded from Total Retail; contributes to All Channels.
  • <b>Home Depot</b> sell-in includes HD Canada (~$20-50K/wk); HD CA sell-through and promo not available (under-reported asymmetry, accepted).
  • <b>TikTok Shop</b> sell-in, sell-through, promo include both US and Mexico. Ads (GMV Max) is US-only — no MX source.
  • Walmart promo only reliable from week ending 2026-04-26 (Scintilla coverage from WM Sat 2026-04-18).
  • Amazon NTB only from week ending 2026-04-05 (SB NTB v3 API 60-day window); 1P only (SB NTB 3P dropped).
  • Amazon Ads on-site = SP 1P only (SP 3P dropped). DSP / SD NTB not included. Walmart / HD / Other Omni Ads — no actual feed yet.
  • New Customers definitions vary by row: All Channels = total first-binders; Total Retail = non-Shopify first-binders; Amazon = SB NTB orders (1P). Not summable across rows.
</div>
</body></html>""")

    with open(dst_path, 'w', encoding='utf-8') as f:
        f.write(''.join(parts))


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Build retail dashboard Excel + HTML from SQL CSV.")
    parser.add_argument('csv', help='Path to SQL output CSV')
    parser.add_argument('--snapshot-date',
                        help='YYYY-MM-DD for output naming + HTML header (default: parse from CSV filename, or today)')
    parser.add_argument('--output-dir',
                        help='Override output dir (default: outputs/<snapshot-date>/)')
    args = parser.parse_args()

    csv_path = os.path.abspath(args.csv)
    if not os.path.exists(csv_path):
        sys.exit(f"CSV not found: {csv_path}")

    snapshot_date = args.snapshot_date or parse_snapshot_date_from_csv_path(csv_path) or date.today().isoformat()

    output_dir = Path(args.output_dir) if args.output_dir else (RESULTS_BASE / snapshot_date)
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / f"Retail Business Dashboard_{snapshot_date}.xlsx"
    html_path = output_dir / f"Retail Business Dashboard_{snapshot_date}.html"

    print(f"Loading CSV: {csv_path}")
    data = load_data(csv_path)
    print(f"  Loaded {len(data)} rows.")

    print("Recomputing Total Retail from per-channel rows (safety net)...")
    recompute_total_retail_including_tts(data)

    print("Loading FY26 plans (Marketing + Promo)...")
    marketing_plan = load_marketing_plan()
    promo_plan = load_promo_plan()

    print("Loading Finance Actuals (Monthly)...")
    finance_actuals = load_finance_actuals()

    print(f"Building Excel → {xlsx_path}")
    stats = build_excel(data, str(xlsx_path), marketing_plan, promo_plan, finance_actuals, snapshot_date)
    print(f"  Wrote {stats['values']} values + {stats['formulas']} formulas + {stats.get('plan_cells', 0)} monthly plan cells + {stats.get('finance_cells', 0)} finance cells. Skipped {stats['skipped']} out-of-range weeks.")

    print(f"Building HTML → {html_path}")
    build_html(data, str(html_path), snapshot_date, marketing_plan, promo_plan, finance_actuals)
    print(f"  HTML size: {os.path.getsize(html_path)/1024:.1f} KB")

    print("\nDone.")
    print(f"  Excel: {xlsx_path}")
    print(f"  HTML : {html_path}")


if __name__ == '__main__':
    main()
